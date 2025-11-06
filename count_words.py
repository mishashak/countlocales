import os
import sys
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from tqdm import tqdm
import tempfile
import json
import shutil
from collections import defaultdict, Counter

# 언어 감지 및 자연어 처리 라이브러리
try:
    from langdetect import detect, DetectorFactory
    DetectorFactory.seed = 0  # 재현 가능한 결과를 위해
except ImportError:
    print("Warning: langdetect not installed. Please install with: pip install langdetect")
    detect = None

try:
    from kiwipiepy import Kiwi
    kiwi = Kiwi()
    print("Korean processor: Kiwi loaded successfully")
except (ImportError, Exception) as e:
    print(f"Warning: kiwipiepy not available ({e}). Korean text will use basic split().")
    kiwi = None

try:
    import spacy
    # 다양한 언어 모델 로드
    nlp_models = {}
    
    # 지원하는 spacy 모델들 (현재 버전에서 사용 가능한 것들만)
    spacy_models = {
        'en': 'en_core_web_sm',
        'es': 'es_core_news_sm', 
        'fr': 'fr_core_news_sm',
        'de': 'de_core_news_sm',
        'pt': 'pt_core_news_sm',
        'it': 'it_core_news_sm',
        'ru': 'ru_core_news_sm'
        # tr, vi, th, id 모델들은 현재 spaCy 버전에서 지원되지 않음
    }
    
    for lang_code, model_name in spacy_models.items():
        try:
            # PyInstaller 환경에서 모델 경로 찾기
            import sys
            import os
            
            if getattr(sys, 'frozen', False):
                # 실행 파일 환경
                base_path = sys._MEIPASS
                model_path = os.path.join(base_path, 'spacy_models', model_name)
                
                # 모델 경로 확인 및 로드
                if os.path.exists(model_path):
                    # config.cfg 파일이 있는지 확인
                    config_path = os.path.join(model_path, f'{model_name}-3.8.0', 'config.cfg')
                    if os.path.exists(config_path):
                        nlp_models[lang_code] = spacy.load(model_path)
                        print(f"Loaded {lang_code} model from bundled path: {model_path}")
                    else:
                        # 전체 모델 디렉토리에서 찾기
                        for root, dirs, files in os.walk(model_path):
                            if 'config.cfg' in files:
                                actual_model_path = root
                                nlp_models[lang_code] = spacy.load(actual_model_path)
                                print(f"Loaded {lang_code} model from: {actual_model_path}")
                                break
                        else:
                            raise Exception(f"config.cfg not found in {model_path}")
                else:
                    # 기본 경로로 시도
                    nlp_models[lang_code] = spacy.load(model_name)
                    print(f"Loaded {lang_code} model: {model_name}")
            else:
                # 개발 환경
                nlp_models[lang_code] = spacy.load(model_name)
                print(f"Loaded {lang_code} model: {model_name}")
        except Exception as e:
            print(f"Warning: {model_name} not available ({e}). Will use basic split().")
        
except ImportError:
    print("Warning: spacy not installed. Please install with: pip install spacy")
    nlp_models = {}

try:
    import jieba
except ImportError:
    print("Warning: jieba not installed. Please install with: pip install jieba")
    jieba = None

try:
    import stanza
    # 일본어 모델 다운로드 및 로드
    try:
        nlp_ja = stanza.Pipeline('ja', verbose=False)
        print("Japanese processor: Stanza loaded successfully")
    except:
        # 모델이 없으면 다운로드 시도
        stanza.download('ja', verbose=False)
        nlp_ja = stanza.Pipeline('ja', verbose=False)
        print("Japanese processor: Stanza loaded successfully")
except (ImportError, Exception) as e:
    print(f"Warning: stanza not available ({e}). Japanese text will use basic split().")
    nlp_ja = None

from translations import t

# 지원 언어 매핑 (langdetect 코드 -> 표시명)
LANGUAGE_MAPPING = {
    'ko': 'Korean',
    'en': 'English', 
    'zh-cn': 'Simplified_Chinese',
    'zh-tw': 'Traditional_Chinese',
    'ja': 'Japanese',
    'vi': 'Vietnamese',
    'th': 'Thai',
    'id': 'Indonesian',
    'ru': 'Russian',
    'es': 'Spanish',
    'pt': 'Portuguese',
    'tr': 'Turkish',
    'fr': 'French',
    'it': 'Italian',
    'de': 'German'
}

# HTML/XML 및 특수 텍스트 패턴
SPECIAL_PATTERNS = {
    'html_xml': re.compile(r'(</?[^<>]*?>)'),
    'brackets': re.compile(r'(\{[^{}]+\})'),
    'newlines': re.compile(r'(\\n)'),
    'file_paths': re.compile(r'([a-zA-Z]:\\[^ ]+|/[^ ]+)')
}

def detect_language(text):
    """텍스트의 언어를 감지"""
    if not detect:
        return 'unknown'
    
    try:
        # 너무 짧은 텍스트는 감지하지 않음
        if len(text.strip()) < 3:
            return 'unknown'
        return detect(text)
    except:
        return 'unknown'

def process_text_by_language(text, language):
    """언어별로 텍스트를 단어로 분리"""
    if not text or pd.isna(text) or str(text).strip() == '':
        return []
    
    text = str(text).strip()
    
    # 특수 패턴 제거
    clean_text = text
    for pattern_name, pattern in SPECIAL_PATTERNS.items():
        clean_text = pattern.sub(' ', clean_text)
    
    # 전처리: 구두점과 하이픈 제거, 숫자/날짜/버전 패턴 보존
    clean_text = preprocess_text(clean_text)
    
    # 공백으로 분리하여 기본 단어 추출
    words = []
    
    if language == 'ko' and kiwi:
        # 한국어: Kiwi 사용
        try:
            tokens = kiwi.tokenize(clean_text)
            words = [token.form for token in tokens if token.form.strip()]
        except:
            words = clean_text.split()
    
    elif language in nlp_models:
        # spaCy 지원 언어들: 영어, 스페인어, 프랑스어, 독일어, 포르투갈어, 이탈리아어, 러시아어, 터키어, 베트남어, 태국어, 인도네시아어
        try:
            doc = nlp_models[language](clean_text)
            words = [token.text for token in doc if not token.is_space and token.text.strip()]
        except:
            words = clean_text.split()
    
    elif language in ['zh-cn', 'zh-tw'] and jieba:
        # 중국어(간체/번체): jieba 사용
        try:
            words = list(jieba.cut(clean_text))
            words = [word for word in words if word.strip()]
        except:
            words = clean_text.split()
    
    elif language == 'ja' and nlp_ja:
        # 일본어: Stanza 사용
        try:
            doc = nlp_ja(clean_text)
            words = []
            for sent in doc.sentences:
                for token in sent.tokens:
                    words.append(token.text)
        except:
            words = clean_text.split()
    
    else:
        # 기타 언어: 기본 split() 사용
        words = clean_text.split()
    
    return [word for word in words if word.strip()]

def extract_special_patterns(text):
    """특수 패턴들을 추출하여 카테고리별로 분류"""
    if not text or pd.isna(text) or str(text).strip() == '':
        return {}
    
    text = str(text)
    pattern_counts = {}
    
    for pattern_name, pattern in SPECIAL_PATTERNS.items():
        matches = pattern.findall(text)
        pattern_counts[pattern_name] = len(matches)
    
    return pattern_counts

def preprocess_text(text):
    """텍스트 전처리: 구두점/하이픈 제거, 숫자/날짜/버전 패턴 보존"""
    import re
    
    # 숫자, 날짜, 버전 패턴들을 먼저 보호 (임시 플레이스홀더로 교체)
    protected_patterns = []
    
    # 버전 패턴 (예: 1.0.4, 2.1.3.5, v1.2.3)
    version_pattern = r'\b(?:v)?\d+(?:\.\d+){1,3}\b'
    for i, match in enumerate(re.finditer(version_pattern, text)):
        placeholder = f"__VERSION_{i}__"
        protected_patterns.append((placeholder, match.group()))
        text = text.replace(match.group(), placeholder, 1)
    
    # 날짜 패턴 (예: 2024-01-15, 15/01/2024, 2024.01.15)
    date_patterns = [
        r'\b\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\b',  # YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD
        r'\b\d{1,2}[-/.]\d{1,2}[-/.]\d{4}\b',  # MM-DD-YYYY, MM/DD/YYYY, MM.DD.YYYY
        r'\b\d{1,2}[-/.]\d{1,2}[-/.]\d{2}\b'   # MM-DD-YY, MM/DD/YY, MM.DD.YY
    ]
    for pattern in date_patterns:
        for i, match in enumerate(re.finditer(pattern, text)):
            placeholder = f"__DATE_{len(protected_patterns)}__"
            protected_patterns.append((placeholder, match.group()))
            text = text.replace(match.group(), placeholder, 1)
    
    # 시간 패턴 (예: 14:30, 2:30:45)
    time_pattern = r'\b\d{1,2}:\d{2}(?::\d{2})?\b'
    for i, match in enumerate(re.finditer(time_pattern, text)):
        placeholder = f"__TIME_{len(protected_patterns)}__"
        protected_patterns.append((placeholder, match.group()))
        text = text.replace(match.group(), placeholder, 1)
    
    # 숫자 패턴 (정수, 소수, 퍼센트, 통화)
    number_patterns = [
        r'\b\d+\.\d+\b',      # 소수 (예: 3.14, 123.45)
        r'\b\d+%\b',          # 퍼센트 (예: 50%, 100%)
        r'\b\d+[km]?\b',      # 숫자 + 단위 (예: 100, 5k, 2m)
        r'\$\d+(?:\.\d{2})?\b',  # 통화 (예: $100, $99.99)
        r'\b\d+\b'            # 정수 (예: 123, 456)
    ]
    for pattern in number_patterns:
        for i, match in enumerate(re.finditer(pattern, text)):
            placeholder = f"__NUMBER_{len(protected_patterns)}__"
            protected_patterns.append((placeholder, match.group()))
            text = text.replace(match.group(), placeholder, 1)
    
    # 하이픈 제거 (단어 합치기)
    text = re.sub(r'-', '', text)
    
    # 구두점 제거 (공백으로 대체)  
    text = re.sub(r'[^\w\s]', ' ', text)
    
    # 연속된 공백을 하나로 정리
    text = re.sub(r'\s+', ' ', text).strip()
    
    # 보호된 패턴들을 원래 값으로 복원
    for placeholder, original in protected_patterns:
        text = text.replace(placeholder, original)
    
    return text

def detect_column_language(df, column_index):
    """특정 열의 모든 셀을 분석하여 가장 많이 나타나는 언어를 반환"""
    if not detect:
        return 'unknown'
    
    language_votes = []
    
    for row in range(df.shape[0]):
        cell_value = df.iat[row, column_index]
        if pd.isna(cell_value) or str(cell_value).strip() == '':
            continue
        
        text = str(cell_value)
        if len(text.strip()) >= 3:  # 최소 길이 체크
            detected_lang = detect_language(text)
            if detected_lang != 'unknown':
                language_votes.append(detected_lang)
    
    if not language_votes:
        return 'unknown'
    
    # 가장 많이 나타나는 언어 반환
    return Counter(language_votes).most_common(1)[0][0]

def count_words_in_text(text, language):
    """텍스트에서 단어 수를 계산 (중복 포함)"""
    words = process_text_by_language(text, language)
    return len(words)

def count_unique_words_in_text(text, language):
    """텍스트에서 고유 단어 수를 계산"""
    words = process_text_by_language(text, language)
    return len(set(words))

# 임시 파일 관리를 위한 클래스 (단어용)
class TempWordManager:
    def __init__(self, base_dir):
        self.base_dir = base_dir
        self.temp_dir = tempfile.mkdtemp(dir=base_dir)
        self.temp_files = defaultdict(list)
        self.current_sets = defaultdict(set)
        self.set_size_limit = 100000

    def add_words(self, category, words):
        for word in words:
            if word not in self.current_sets[category]:
                self.current_sets[category].add(word)
                if len(self.current_sets[category]) >= self.set_size_limit:
                    self._save_to_temp_file(category)

    def _save_to_temp_file(self, category):
        if not self.current_sets[category]:
            return

        temp_file = os.path.join(self.temp_dir, f"{category}_{len(self.temp_files[category])}.json")
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(list(self.current_sets[category]), f, ensure_ascii=False)
        
        self.temp_files[category].append(temp_file)
        self.current_sets[category].clear()

    def get_all_unique_words(self, category):
        all_words = set()
        
        # 현재 메모리에 있는 단어 추가
        all_words.update(self.current_sets[category])
        
        # 임시 파일에서 단어 로드
        for temp_file in self.temp_files[category]:
            with open(temp_file, 'r', encoding='utf-8') as f:
                all_words.update(json.load(f))
        
        return all_words

    def cleanup(self):
        shutil.rmtree(self.temp_dir)

def analyze_sheet_for_words(df):
    """시트를 분석하여 단어 수를 계산"""
    # 먼저 각 열의 언어를 감지
    column_languages = {}
    for col in range(df.shape[1]):
        column_languages[col] = detect_column_language(df, col)
    
    # 전체 카테고리 (언어 + 특수 패턴)
    all_categories = set()
    for lang_code in column_languages.values():
        if lang_code != 'unknown':
            # langdetect 코드를 표시명으로 변환
            display_name = LANGUAGE_MAPPING.get(lang_code, lang_code)
            all_categories.add(display_name)
    
    all_categories.update(['html_xml', 'brackets', 'newlines', 'file_paths'])
    
    total_counts = {category: 0 for category in all_categories}
    column_counts = {col: {category: 0 for category in all_categories} for col in range(df.shape[1])}
    
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell_value = df.iat[r, c]
            if pd.isna(cell_value) or str(cell_value).strip() == '':
                continue
            
            text = str(cell_value)
            col_lang = column_languages[c]
            
            if col_lang != 'unknown':
                # langdetect 코드를 표시명으로 변환
                display_name = LANGUAGE_MAPPING.get(col_lang, col_lang)
                # 단어 수 계산
                word_count = count_words_in_text(text, col_lang)
                total_counts[display_name] += word_count
                column_counts[c][display_name] += word_count
            
            # 특수 패턴 카운트
            special_patterns = extract_special_patterns(text)
            for pattern_name, count in special_patterns.items():
                if count > 0:
                    total_counts[pattern_name] += count
                    column_counts[c][pattern_name] += count
    
    # 유효한 열만 필터링
    valid_columns = []
    empty_col_count = 0
    for col in range(df.shape[1]):
        col_total = sum(column_counts[col].values())
        if col_total > 0:
            valid_columns.append(col)
            empty_col_count = 0
        else:
            empty_col_count += 1
            if empty_col_count >= 20:
                break
    
    return total_counts, column_counts, valid_columns, column_languages

def get_unique_words_per_column(df, column_languages):
    """각 열별로 고유 단어 수를 계산"""
    unique_counts = {}
    for col in range(df.shape[1]):
        col_lang = column_languages.get(col, 'unknown')
        
        # 전체 카테고리 생성
        all_categories = set()
        for lang_code in column_languages.values():
            if lang_code != 'unknown':
                display_name = LANGUAGE_MAPPING.get(lang_code, lang_code)
                all_categories.add(display_name)
        all_categories.update(['html_xml', 'brackets', 'newlines', 'file_paths'])
        
        col_counts = {category: 0 for category in all_categories}
        
        # 해당 열의 모든 값을 가져옴
        column_values = df.iloc[:, col].dropna().astype(str)
        unique_texts = column_values.unique()
        
        # 각 고유 값에 대해 단어 수를 세고 합산
        for value in unique_texts:
            if col_lang != 'unknown':
                display_name = LANGUAGE_MAPPING.get(col_lang, col_lang)
                unique_word_count = count_unique_words_in_text(value, col_lang)
                col_counts[display_name] += unique_word_count
            
            # 특수 패턴 카운트
            special_patterns = extract_special_patterns(value)
            for pattern_name, count in special_patterns.items():
                col_counts[pattern_name] += count
        
        unique_counts[col] = col_counts
    
    return unique_counts

def get_cell_addresses_for_words(df, column_languages):
    """단어 수 분석용 셀 주소 추출"""
    cell_addresses = {}
    
    # 전체 카테고리 생성
    all_categories = set()
    for lang_code in column_languages.values():
        if lang_code != 'unknown':
            display_name = LANGUAGE_MAPPING.get(lang_code, lang_code)
            all_categories.add(display_name)
    all_categories.update(['html_xml', 'brackets', 'newlines', 'file_paths'])
    
    # 각 카테고리별로 셀 주소 딕셔너리 초기화
    for category in all_categories:
        cell_addresses[category] = {col: [] for col in range(df.shape[1])}
    
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell_value = df.iat[r, c]
            if pd.isna(cell_value) or str(cell_value).strip() == '':
                continue
                
            text = str(cell_value)
            col_lang = column_languages.get(c, 'unknown')
            
            if col_lang != 'unknown':
                # langdetect 코드를 표시명으로 변환
                display_name = LANGUAGE_MAPPING.get(col_lang, col_lang)
                # 단어가 있는지 확인
                word_count = count_words_in_text(text, col_lang)
                if word_count > 0:
                    cell_address = f"{get_column_letter(c+1)}{r+1}"
                    cell_addresses[display_name][c].append(cell_address)
            
            # 특수 패턴 확인
            special_patterns = extract_special_patterns(text)
            for pattern_name, count in special_patterns.items():
                if count > 0:
                    cell_address = f"{get_column_letter(c+1)}{r+1}"
                    cell_addresses[pattern_name][c].append(cell_address)
    
    # 각 카테고리별로 셀 주소 정렬
    for category in all_categories:
        for col in range(df.shape[1]):
            cell_addresses[category][col].sort(key=lambda x: (x[0], int(x[1:])))
    
    return cell_addresses

def count_cells_by_category_for_words(df, column_languages):
    """단어 수 분석용 카테고리별 셀 개수 계산"""
    cell_counts = {}
    
    # 전체 카테고리 생성
    all_categories = set()
    for lang_code in column_languages.values():
        if lang_code != 'unknown':
            display_name = LANGUAGE_MAPPING.get(lang_code, lang_code)
            all_categories.add(display_name)
    all_categories.update(['html_xml', 'brackets', 'newlines', 'file_paths'])
    
    # 각 카테고리별로 셀 개수 딕셔너리 초기화
    for category in all_categories:
        cell_counts[category] = {col: 0 for col in range(df.shape[1])}
    
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell_value = df.iat[r, c]
            if pd.isna(cell_value) or str(cell_value).strip() == '':
                continue
                
            text = str(cell_value)
            col_lang = column_languages.get(c, 'unknown')
            
            if col_lang != 'unknown':
                # langdetect 코드를 표시명으로 변환
                display_name = LANGUAGE_MAPPING.get(col_lang, col_lang)
                # 단어가 있는지 확인
                word_count = count_words_in_text(text, col_lang)
                if word_count > 0:
                    cell_counts[display_name][c] += 1
            
            # 특수 패턴 확인
            special_patterns = extract_special_patterns(text)
            for pattern_name, count in special_patterns.items():
                if count > 0:
                    cell_counts[pattern_name][c] += 1
    
    return cell_counts

def adjust_column_widths(sheet):
    """열 너비 조정"""
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        
        # Words_cell_address 시트에서 데이터 열만 너비를 10으로 고정
        if sheet.title == 'Words_cell_address' and column >= 'G':
            sheet.column_dimensions[column].width = 10
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=False, shrink_to_fit=True)
            continue
            
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

def main(current_language='ko'):
    """메인 함수"""
    # exe 파일이 실행된 경로를 기준으로 설정
    folder_path = os.path.dirname(os.path.abspath(sys.executable)) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    print(f"{t('UI_006', current_language)}: {folder_path}")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_name = f"WORD_COUNT_REPORT_{timestamp}.xlsx"
    report_path = os.path.join(folder_path, report_name)
    print(f"{t('UI_007', current_language)}: {report_path}")

    report_wb = Workbook()
    
    # 6개의 시트 생성 (count_chars와 동일한 구조)
    report_ws_real = report_wb.active
    report_ws_real.title = 'Words_real'
    
    report_ws_unique_for_sheet = report_wb.create_sheet('Words_unique_for_Sheet')
    report_ws_unique_for_folder = report_wb.create_sheet('Words_unique_for_Folder')
    report_ws_cell_address = report_wb.create_sheet('Words_cell_address')
    report_ws_cells = report_wb.create_sheet('Words_cells')
    
    # 임시 파일 매니저 초기화
    temp_manager = TempWordManager(folder_path)

    # 하위 폴더를 포함한 모든 엑셀 파일 수집
    files_to_process = []
    for root, dirs, files in os.walk(folder_path):
        if '__pycache__' in dirs:
            dirs.remove('__pycache__')
        if '.git' in dirs:
            dirs.remove('.git')
            
        for file in files:
            if file.endswith(('.xlsx', '.xlsm', '.csv')) and "REPORT_" not in file and not file.startswith('~$'):
                rel_path = os.path.relpath(os.path.join(root, file), folder_path)
                files_to_process.append((rel_path, file))

    print(f"{t('UI_008', current_language)}: {[f[1] for f in files_to_process]}")
    print(t('UI_009', current_language).format(len(files_to_process)))
    print(t('UI_010', current_language))

    # 전체 열을 추적하기 위한 변수
    all_columns = set()
    all_categories = set()
    processed_files = 0

    data_rows_real = []
    data_rows_unique_for_sheet = []
    data_rows_unique_for_folder = []
    data_rows_cell_address = []
    data_rows_cells = []

    # 콘솔 출력이 가능한지 확인
    has_console = hasattr(sys.stdout, 'write') and sys.stdout is not None
    
    for rel_path, file_name in tqdm(files_to_process, desc="processing files", disable=not has_console):
        try:
            print(f"\n{t('UI_011', current_language)}: {file_name}")
            file_path = os.path.join(folder_path, rel_path)
            xls = pd.ExcelFile(file_path)

            for sheet_name in xls.sheet_names:
                print(f"{t('UI_012', current_language)}: {sheet_name}")
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                
                total_counts, column_counts, valid_columns, column_languages = analyze_sheet_for_words(df)
                unique_counts = get_unique_words_per_column(df, column_languages)
                cell_addresses = get_cell_addresses_for_words(df, column_languages)
                cell_counts = count_cells_by_category_for_words(df, column_languages)

                # 유효한 열을 전체 열 목록에 추가
                for col in valid_columns:
                    all_columns.add(col)

                # 전체 카테고리 업데이트
                all_categories.update(total_counts.keys())

                # 고유한 단어 수집 (폴더 전체 기준)
                for r in range(df.shape[0]):
                    for c in range(df.shape[1]):
                        cell_value = df.iat[r, c]
                        if pd.isna(cell_value) or str(cell_value).strip() == '':
                            continue
                        
                        text = str(cell_value)
                        col_lang = column_languages.get(c, 'unknown')
                        
                        if col_lang != 'unknown':
                            display_name = LANGUAGE_MAPPING.get(col_lang, col_lang)
                            words = process_text_by_language(text, col_lang)
                            temp_manager.add_words(display_name, words)

                # 실제 데이터 처리
                for category in total_counts:
                    if category in ['html_xml', 'brackets', 'newlines', 'file_paths']:
                        emoji = '🔧'  # 특수 패턴용 이모지
                    else:
                        emoji = '🌐'  # 언어용 이모지
                    
                    col_totals = [column_counts[col].get(category, 0) for col in valid_columns]
                    total = total_counts[category]
                    sum_col_totals = sum(col_totals)
                    
                    if sum_col_totals != total:
                        status = f"Error: Total words({total}) and column totals({sum_col_totals}) do not match"
                    else:
                        status = "Normal"

                    row_data = [rel_path, file_name, sheet_name, status, emoji, category, total] + col_totals
                    data_rows_real.append(row_data)

                # 고유 값 데이터 처리
                for category in total_counts:
                    if category in ['html_xml', 'brackets', 'newlines', 'file_paths']:
                        emoji = '🔧'
                    else:
                        emoji = '🌐'
                    
                    unique_col_totals = [unique_counts[col].get(category, 0) for col in valid_columns]
                    total_unique = sum(unique_col_totals)
                    row_data = [rel_path, file_name, sheet_name, "Normal", emoji, category, total_unique] + unique_col_totals
                    data_rows_unique_for_sheet.append(row_data)
                
                # 셀 주소 데이터 처리
                for category in total_counts:
                    if category in ['html_xml', 'brackets', 'newlines', 'file_paths']:
                        emoji = '🔧'
                    else:
                        emoji = '🌐'
                    
                    cell_col_addresses = [', '.join(cell_addresses[category][col]) for col in valid_columns]
                    total_cells = sum(len(cell_addresses[category][col]) for col in valid_columns)
                    row_data = [rel_path, file_name, sheet_name, "Normal", emoji, category, total_cells] + cell_col_addresses
                    data_rows_cell_address.append(row_data)
                
                # 셀 갯수 데이터 처리
                for category in total_counts:
                    if category in ['html_xml', 'brackets', 'newlines', 'file_paths']:
                        emoji = '🔧'
                    else:
                        emoji = '🌐'
                    
                    cell_col_counts = [cell_counts[category][col] for col in valid_columns]
                    total_cells = sum(cell_col_counts)
                    row_data = [rel_path, file_name, sheet_name, "Normal", emoji, category, total_cells] + cell_col_counts
                    data_rows_cells.append(row_data)

            processed_files += 1
            print(t('UI_013', current_language).format(f"{processed_files}/{len(files_to_process)}"))

        except Exception as e:
            print(f"{t('UI_017', current_language)}: {file_name} {t('UI_018', current_language)}: {e}")
            continue

    print(f"\n{t('UI_014', current_language)}")
    
    # 시트에 데이터 추가
    sorted_columns = sorted(all_columns)
    column_headers = [f"Col {get_column_letter(col+1)}" for col in sorted_columns]
    headers = ['Path', 'FileName', 'SheetName', 'Status', '🏳️', 'Category', 'TotalWords'] + column_headers
    
    # Words_real 시트에 데이터 추가
    report_ws_real.append(headers)
    for row in data_rows_real:
        report_ws_real.append(row)
    adjust_column_widths(report_ws_real)

    # Words_unique_for_Sheet 시트에 데이터 추가
    report_ws_unique_for_sheet.append(headers)
    for row in data_rows_unique_for_sheet:
        report_ws_unique_for_sheet.append(row)
    adjust_column_widths(report_ws_unique_for_sheet)
    
    # Words_unique_for_Folder 시트에 데이터 추가
    report_ws_unique_for_folder.append(headers)
    for category in sorted(all_categories):
        if category in ['html_xml', 'brackets', 'newlines', 'file_paths']:
            emoji = '🔧'
        else:
            emoji = '🌐'
        
        unique_words = temp_manager.get_all_unique_words(category)
        total_unique_words = len(unique_words)
        row_data = ['ALL', 'ALL', 'ALL', 'Normal', emoji, category, total_unique_words] + [0] * len(sorted_columns)
        report_ws_unique_for_folder.append(row_data)
    adjust_column_widths(report_ws_unique_for_folder)
    
    # Words_cell_address 시트에 데이터 추가
    cell_address_headers = headers.copy()
    cell_address_headers[5] = 'TotalCells'  # F1 셀의 헤더 변경
    report_ws_cell_address.append(cell_address_headers)
    for row in data_rows_cell_address:
        report_ws_cell_address.append(row)
    adjust_column_widths(report_ws_cell_address)
    
    # Words_cells 시트에 데이터 추가
    cells_headers = headers.copy()
    cells_headers[5] = 'TotalCells'  # F1 셀의 헤더 변경
    report_ws_cells.append(cells_headers)
    for row in data_rows_cells:
        report_ws_cells.append(row)
    adjust_column_widths(report_ws_cells)

    # 임시 파일 정리
    temp_manager.cleanup()

    report_wb.save(report_path)
    print(f"{t('UI_015', current_language)}: {report_path}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error: {e}")
        input("Press any key to continue...")
