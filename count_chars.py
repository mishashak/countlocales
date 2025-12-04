import os
import sys
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from tqdm import tqdm
import tempfile
import json
import shutil
from translations import t

# 언어별 정규 표현식 패턴 정의 및 간결한 이름
PATTERNS = {
    'Korean': (re.compile(r'[가-힣]'), '🇰🇷'),
    'Alphabet': (re.compile(r'[A-Za-zÀ-ÿ]'), '🇺🇸'),
    'Number': (re.compile(r'[0-9]'), '🔢'),
    'Chinese': (re.compile(r'[\u4E00-\u9FFF]'), '🇨🇳'),
    'Japanese': (re.compile(r'[\u30A0-\u30FF\u3040-\u309F]'), '🇯🇵'),
    'Thai': (re.compile(r'[\u0E00-\u0E7F]'), '🇹🇭'),
    'Russian': (re.compile(r'[\u0400-\u04FF]'), '🇷🇺'),  # 키릴 문자 추가
    'Special': (re.compile(r'[^\w\s]'), '🔣')
}

# 임시 파일 관리를 위한 클래스
class TempFileManager:
    def __init__(self, base_dir):
        self.base_dir = base_dir
        self.temp_dir = tempfile.mkdtemp(dir=base_dir)
        self.temp_files = {lang: [] for lang in PATTERNS}
        self.current_sets = {lang: set() for lang in PATTERNS}
        self.set_size_limit = 100000  # 각 언어별 집합 크기 제한

    def add_text(self, lang, text):
        if text in self.current_sets[lang]:
            return
        
        self.current_sets[lang].add(text)
        if len(self.current_sets[lang]) >= self.set_size_limit:
            self._save_to_temp_file(lang)

    def _save_to_temp_file(self, lang):
        if not self.current_sets[lang]:
            return

        temp_file = os.path.join(self.temp_dir, f"{lang}_{len(self.temp_files[lang])}.json")
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(list(self.current_sets[lang]), f, ensure_ascii=False)
        
        self.temp_files[lang].append(temp_file)
        self.current_sets[lang].clear()

    def get_all_unique_texts(self, lang):
        all_texts = set()
        
        # 현재 메모리에 있는 텍스트 추가
        all_texts.update(self.current_sets[lang])
        
        # 임시 파일에서 텍스트 로드
        for temp_file in self.temp_files[lang]:
            with open(temp_file, 'r', encoding='utf-8') as f:
                all_texts.update(json.load(f))
        
        return all_texts

    def get_total_chars(self, lang):
        total_chars = 0
        unique_texts = self.get_all_unique_texts(lang)
        for text in unique_texts:
            counts = count_characters(text)
            total_chars += counts[lang]
        return total_chars

    def cleanup(self):
        """임시 디렉토리 정리 (Windows 액세스 거부 오류 처리)"""
        import time
        import stat
        
        def handle_remove_readonly(func, path, exc):
            """읽기 전용 파일 삭제를 위한 핸들러"""
            os.chmod(path, stat.S_IWRITE)
            func(path)
        
        max_retries = 3
        retry_delay = 0.5
        
        for attempt in range(max_retries):
            try:
                if os.path.exists(self.temp_dir):
                    shutil.rmtree(self.temp_dir, onerror=handle_remove_readonly)
                break
            except (PermissionError, OSError) as e:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    retry_delay *= 2  # 지수 백오프
                else:
                    # 최종 시도 실패 시 경고만 출력하고 계속 진행
                    print(f"Warning: Could not delete temporary directory {self.temp_dir}: {e}")
                    print(f"Please manually delete it if needed.")

def count_characters(text):
    counts = {lang: 0 for lang in PATTERNS}
    for lang, (pattern, _) in PATTERNS.items():
        matches = pattern.findall(text)
        counts[lang] += len(matches)
    return counts

def determine_primary_language(counts):
    non_special_counts = {lang: count for lang, count in counts.items() if lang not in ['Special']}
    non_english_counts = {lang: count for lang, count in non_special_counts.items() if lang != ['Alphabet']}

    if sum(non_special_counts.values()) == 0:
        # 특수 문자와 숫자만 있는 경우
        primary_lang = max(non_special_counts, key=non_special_counts.get)
    elif sum(non_english_counts.values()) > 0:
        # 영어 외에 다른 언어가 있는 경우
        primary_lang = max(non_english_counts, key=non_english_counts.get)
    else:
        # 영어만 있는 경우
        primary_lang = 'Alphabet'
    
    return primary_lang

def analyze_sheet(df):
    total_counts = {lang: 0 for lang in PATTERNS}
    column_counts = {col: {lang: 0 for lang in PATTERNS} for col in range(df.shape[1])}

    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell_value = df.iat[r, c]
            if pd.isna(cell_value) or str(cell_value).strip() == '':
                continue  # 빈 셀은 무시
            text = str(cell_value)
            counts = count_characters(text)
            
            # 모든 언어의 글자 수를 더함
            for lang in PATTERNS:
                total_counts[lang] += counts[lang]
                column_counts[c][lang] += counts[lang]

    # 유효한 열만 필터링
    valid_columns = []
    empty_col_count = 0
    for col in range(df.shape[1]):
        col_total = sum(column_counts[col].values())
        if col_total > 0:
            valid_columns.append(col)
            empty_col_count = 0
        else:
            empty_col_count += 1
            if empty_col_count >= 20:  # 빈 열이 20개 이상 연속될 경우 중단
                break

    return total_counts, column_counts, valid_columns

def adjust_column_widths(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column name
        
        # Summary_cell_address 시트에서 데이터 열(A열, B열 등)만 너비를 10으로 고정하고 셀에 맞춤 설정
        if sheet.title == 'Summary_cell_address' and column >= 'G':  # G열부터 시작하는 데이터 열
            sheet.column_dimensions[column].width = 10
            # 해당 열의 모든 셀에 셀에 맞춤 설정 적용 (줄바꿈 없음)
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=False, shrink_to_fit=True)
            continue
            
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

def get_unique_values_per_column(df):
    unique_counts = {}
    for col in range(df.shape[1]):
        # 해당 열의 모든 값을 가져옴
        column_values = df.iloc[:, col].dropna().astype(str)
        # 중복 제거
        unique_values = column_values.unique()
        # 각 고유 값에 대해 글자 수를 세고 합산
        col_counts = {lang: 0 for lang in PATTERNS}
        for value in unique_values:
            counts = count_characters(value)
            for lang in PATTERNS:
                col_counts[lang] += counts[lang]
        unique_counts[col] = col_counts
    return unique_counts

def get_cell_addresses(df):
    cell_addresses = {lang: {col: [] for col in range(df.shape[1])} for lang in PATTERNS}
    
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell_value = df.iat[r, c]
            if pd.isna(cell_value) or str(cell_value).strip() == '':
                continue  # 빈 셀은 무시
                
            text = str(cell_value)
            for lang, (pattern, _) in PATTERNS.items():
                if pattern.search(text):
                    cell_address = f"{get_column_letter(c+1)}{r+1}"
                    cell_addresses[lang][c].append(cell_address)
    
    # 각 언어별로 셀 주소 정렬
    for lang in PATTERNS:
        for col in range(df.shape[1]):
            cell_addresses[lang][col].sort(key=lambda x: (x[0], int(x[1:])))
    
    return cell_addresses

def count_cells_by_language(df):
    cell_counts = {lang: {col: 0 for col in range(df.shape[1])} for lang in PATTERNS}
    
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell_value = df.iat[r, c]
            if pd.isna(cell_value) or str(cell_value).strip() == '':
                continue  # 빈 셀은 무시
                
            text = str(cell_value)
            for lang, (pattern, _) in PATTERNS.items():
                if pattern.search(text):
                    cell_counts[lang][c] += 1
    
    return cell_counts

def main(current_language='ko'):
    # exe 파일이 실행된 경로를 기준으로 설정
    folder_path = os.path.dirname(os.path.abspath(sys.executable)) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
    print(f"{t('UI_006', current_language)}: {folder_path}")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_name = f"CHAR_COUNT_REPORT_{timestamp}.xlsx"
    report_path = os.path.join(folder_path, report_name)
    print(f"{t('UI_007', current_language)}: {report_path}")

    report_wb = Workbook()
    
    # Summary_real 시트 생성
    report_ws_real = report_wb.active
    report_ws_real.title = 'Summary_real'
    
    # Summary_unique_for_Sheet 시트 생성
    report_ws_unique_for_sheet = report_wb.create_sheet('Summary_unique_for_Sheet')
    
    # Summary_unique_for_Folder 시트 생성
    report_ws_unique_for_folder = report_wb.create_sheet('Summary_unique_for_Folder')
    
    # Summary_cell_address 시트 생성
    report_ws_cell_address = report_wb.create_sheet('Summary_cell_address')
    
    # Summary_cells 시트 생성
    report_ws_cells = report_wb.create_sheet('Summary_cells')

    # 임시 파일 매니저 초기화
    temp_manager = TempFileManager(folder_path)

    # 하위 폴더를 포함한 모든 엑셀 파일 수집
    files_to_process = []
    for root, dirs, files in os.walk(folder_path):
        # 특정 폴더 제외
        if '__pycache__' in dirs:
            dirs.remove('__pycache__')
        if '.git' in dirs:
            dirs.remove('.git')
            
        for file in files:
            if file.endswith(('.xlsx', '.xlsm', '.csv')) and "REPORT_" not in file and not file.startswith('~$'):
                # 상대 경로 계산
                rel_path = os.path.relpath(os.path.join(root, file), folder_path)
                files_to_process.append((rel_path, file))

    print(f"{t('UI_008', current_language)}: {[f[1] for f in files_to_process]}")

    # 전체 열을 추적하기 위한 변수
    all_columns = set()
    processed_files = 0

    print(t('UI_009', current_language).format(len(files_to_process)))
    print(t('UI_010', current_language))

    data_rows_real = []  # 실제 데이터를 저장할 리스트
    data_rows_unique_for_sheet = []  # 고유 값 데이터를 저장할 리스트
    data_rows_cell_address = []  # 셀 주소 데이터를 저장할 리스트
    data_rows_cells = []  # 셀 갯수 데이터를 저장할 리스트

    # 콘솔 출력이 가능한지 확인
    has_console = hasattr(sys.stdout, 'write') and sys.stdout is not None
    
    for rel_path, file_name in tqdm(files_to_process, desc="processing files", disable=not has_console):
        try:
            print(f"\n{t('UI_011', current_language)}: {file_name}")
            file_path = os.path.join(folder_path, rel_path)
            xls = pd.ExcelFile(file_path)

            for sheet_name in xls.sheet_names:
                print(f"{t('UI_012', current_language)}: {sheet_name}")
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                total_counts, column_counts, valid_columns = analyze_sheet(df)
                unique_counts = get_unique_values_per_column(df)
                cell_addresses = get_cell_addresses(df)
                cell_counts = count_cells_by_language(df)

                # 유효한 열을 전체 열 목록에 추가
                for col in valid_columns:
                    all_columns.add(col)

                # 고유한 텍스트 수집 (폴더 전체 기준)
                for r in range(df.shape[0]):
                    for c in range(df.shape[1]):
                        cell_value = df.iat[r, c]
                        if pd.isna(cell_value) or str(cell_value).strip() == '':
                            continue
                        text = str(cell_value)
                        for lang, (pattern, _) in PATTERNS.items():
                            if pattern.search(text):
                                temp_manager.add_text(lang, text)

                # 실제 데이터 처리
                for lang in PATTERNS:
                    emoji = PATTERNS[lang][1]
                    col_totals = [column_counts[col][lang] for col in valid_columns]
                    total = total_counts[lang]
                    sum_col_totals = sum(col_totals)
                    if sum_col_totals != total:
                        status = f"Error: Total characters({total}) and column totals({sum_col_totals}) do not match"
                    else:
                        status = "Normal"

                    row_data = [rel_path, file_name, sheet_name, status, emoji, f"{lang}", total] + col_totals
                    data_rows_real.append(row_data)

                # 고유 값 데이터 처리
                for lang in PATTERNS:
                    emoji = PATTERNS[lang][1]
                    unique_col_totals = [unique_counts[col][lang] for col in valid_columns]
                    total_unique = sum(unique_col_totals)
                    row_data = [rel_path, file_name, sheet_name, "Normal", emoji, f"{lang}", total_unique] + unique_col_totals
                    data_rows_unique_for_sheet.append(row_data)
                
                # 셀 주소 데이터 처리
                for lang in PATTERNS:
                    emoji = PATTERNS[lang][1]
                    cell_col_addresses = [', '.join(cell_addresses[lang][col]) for col in valid_columns]
                    total_cells = sum(len(cell_addresses[lang][col]) for col in valid_columns)
                    row_data = [rel_path, file_name, sheet_name, "Normal", emoji, f"{lang}", total_cells] + cell_col_addresses
                    data_rows_cell_address.append(row_data)
                
                # 셀 갯수 데이터 처리
                for lang in PATTERNS:
                    emoji = PATTERNS[lang][1]
                    cell_col_counts = [cell_counts[lang][col] for col in valid_columns]
                    total_cells = sum(cell_col_counts)
                    row_data = [rel_path, file_name, sheet_name, "Normal", emoji, f"{lang}", total_cells] + cell_col_counts
                    data_rows_cells.append(row_data)

            processed_files += 1
            print(t('UI_013', current_language).format(f"{processed_files}/{len(files_to_process)}"))

        except Exception as e:
            print(f"{t('UI_017', current_language)}: {file_name} {t('UI_018', current_language)}: {e}")
            continue

    print(f"\n{t('UI_014', current_language)}")
    # Summary_real 시트의 헤더 추가
    sorted_columns = sorted(all_columns)
    column_headers = [f"Col {get_column_letter(col+1)}" for col in sorted_columns]
    headers = ['Path', 'FileName', 'SheetName', 'Status', '🏳️', 'Char', 'TotalChars'] + column_headers
    
    # Summary_real 시트에 데이터 추가
    report_ws_real.append(headers)
    for row in data_rows_real:
        report_ws_real.append(row)
    adjust_column_widths(report_ws_real)

    # Summary_unique_for_Sheet 시트에 데이터 추가
    report_ws_unique_for_sheet.append(headers)
    for row in data_rows_unique_for_sheet:
        report_ws_unique_for_sheet.append(row)
    adjust_column_widths(report_ws_unique_for_sheet)
    
    # Summary_unique_for_Folder 시트에 데이터 추가
    report_ws_unique_for_folder.append(headers)
    for lang in PATTERNS:
        emoji = PATTERNS[lang][1]
        # count_characters 함수를 사용하여 글자 수 계산
        total_chars = temp_manager.get_total_chars(lang)
        row_data = ['ALL', 'ALL', 'ALL', 'Normal', emoji, f"{lang}", total_chars] + [0] * len(sorted_columns)
        report_ws_unique_for_folder.append(row_data)
    adjust_column_widths(report_ws_unique_for_folder)
    
    # Summary_cell_address 시트에 데이터 추가
    cell_address_headers = headers.copy()
    cell_address_headers[5] = 'TotalCells'  # F1 셀의 헤더 변경
    report_ws_cell_address.append(cell_address_headers)
    for row in data_rows_cell_address:
        report_ws_cell_address.append(row)
    adjust_column_widths(report_ws_cell_address)
    
    # Summary_cells 시트에 데이터 추가
    cells_headers = headers.copy()
    cells_headers[5] = 'TotalCells'  # F1 셀의 헤더 변경
    report_ws_cells.append(cells_headers)
    for row in data_rows_cells:
        report_ws_cells.append(row)
    adjust_column_widths(report_ws_cells)

    # 임시 파일 정리
    temp_manager.cleanup()

    report_wb.save(report_path)
    print(f"{t('UI_015', current_language)}: {report_path}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error: {e}")
        input("Press any key to continue...")
